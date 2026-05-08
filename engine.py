# ═══════════════════════════════════════════════
# DEEL 1 — IMPORTS EN APP SETUP
# ═══════════════════════════════════════════════

import os
import re
import json
import requests
from datetime import datetime
from bs4 import BeautifulSoup
from flask import Flask, request, jsonify
from flask_cors import CORS
import anthropic

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import redis
    from rq import Queue
    from rq.job import Job
    HAS_REDIS = True
except ImportError:
    HAS_REDIS = False

# Redis verbinding
def get_redis_connection():
    redis_url = os.environ.get("REDIS_URL")
    if not redis_url or not HAS_REDIS:
        return None
    try:
        conn = redis.from_url(redis_url, socket_connect_timeout=5)
        conn.ping()
        return conn
    except Exception as e:
        print(f"REDIS ERROR: {e}", flush=True)
        return None

app = Flask(__name__)
CORS(app, origins=["https://funcify.eu", "https://fullfunc.github.io"])

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CLAUDE_MODEL = "claude-sonnet-4-5"
_cache = {}

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "nl-NL,nl;q=0.9,en;q=0.8",
    "Accept-Encoding": "gzip, deflate, br",
}

# ═══════════════════════════════════════════════
# DEEL 2 — CONSTANTEN
# ═══════════════════════════════════════════════

QUALITY_CERT_LIST = [
    "ifos", "gmp", "iso 22000", "iso22000", "informed sport", "nzvt",
    "creapure", "nsf", "haccp", "usp", "iso 17025", "iso17025",
    "fssc 22000", "brc", "ifs", "traacs", "albion",
]

SUSTAINABILITY_CERT_LIST = [
    "msc", "friend of the sea", "dolphin safe", "rainforest alliance",
    "fair trade", "vegan society", "soil association", "organic",
    "biologisch", "green-e", "golden omega",
]

PRODUCT_TYPE_WEIGHTS = {
    "OMEGA3":       {"review": 0.40, "biology": 0.60},
    "LIPID":        {"review": 0.45, "biology": 0.55},
    "THYROID":      {"review": 0.30, "biology": 0.70},
    "MINERAL":      {"review": 0.40, "biology": 0.60},
    "INFLAMMATION": {"review": 0.35, "biology": 0.65},
    "PROBIOTIC":    {"review": 0.35, "biology": 0.65},
    "STRESS":       {"review": 0.35, "biology": 0.65},
    "VITAMIN":      {"review": 0.40, "biology": 0.60},
    "IRON":         {"review": 0.45, "biology": 0.55},
    "SPORT":        {"review": 0.40, "biology": 0.60},
    "PROTEIN":      {"review": 0.45, "biology": 0.55},
    "BOTANICAL":    {"review": 0.35, "biology": 0.65},
    "DEFAULT":      {"review": 0.40, "biology": 0.60},
}

SEVERITY_ORDER = {"Critical": 0, "Major": 1, "Minor": 2, "Info": 3}

COMPLEXITY_LABELS = {
    "Single": "Enkelvoudig",
    "Multi": "Meervoudig",
    "Complex": "Complex",
}

VASTE_AFSLUITING = """Deze beoordeling is opgesteld vanuit een orthomoleculair perspectief. Dat betekent dat wij niet alleen kijken naar de aanbevolen dagelijkse hoeveelheid, maar naar wat klinisch relevant is op basis van biochemische werkzaamheid, biologische beschikbaarheid en individuele context. Suppletie is een aanvulling op en geen vervanging van gevarieerde voeding, een gezonde leefstijl, voldoende slaap, beweging en stressregulatie. De effectiviteit van elk supplement wordt mede bepaald door jouw persoonlijke gezondheidscontext, medicatiegebruik en voedingsstatus. Dit is geen medisch advies. Gebruik je medicijnen of heb je een medische aandoening? Raadpleeg dan altijd een arts of orthomoleculair specialist. Voor een gepersonaliseerde beoordeling op basis van jouw doel, intake en gezondheidscontext, gebruik het persoonlijke intakeprotocol op het platform."""

CONTEXT_FLAG_RULES = [
    {"id":"CF001","keywords":["pyridoxine","vitamine b6","vitamin b6","p-5-p"],"operator":"gte","threshold":25,"unit":"mg","severity":"Critical","message":"Dit product bevat een hoge B6-dosering. Langdurige inname boven 25 mg per dag kan bij sommige mensen perifere neuropathie veroorzaken. Gebruik niet structureel zonder begeleiding."},
    {"id":"CF002","keywords":["pyridoxine","vitamine b6","p-5-p"],"operator":"gte","threshold":50,"unit":"mg","severity":"Critical","message":"Dit product bevat een zeer hoge B6-dosering. Alleen bij specifieke therapeutische indicatie onder begeleiding."},
    {"id":"CF003","keywords":["retinol","vitamine a","vitamin a"],"operator":"gte","threshold":3000,"unit":"IU","severity":"Critical","message":"Dit product bevat een hoge dosis vitamine A. Accumulatie is mogelijk bij dagelijks gebruik. Niet aanbevolen tijdens zwangerschap."},
    {"id":"CF004","keywords":["retinol","vitamine a"],"operator":"gte","threshold":10000,"unit":"IU","severity":"Critical","message":"Hoge retinol-inname is geassocieerd met aangeboren afwijkingen. Dit product is niet geschikt tijdens zwangerschap."},
    {"id":"CF005","keywords":["jodium","iodine","kaliumjodide","natriumjodide","zeewier","kelp"],"operator":"gte","threshold":150,"unit":"mcg","severity":"Major","message":"Dit product bevat jodium. Bij bestaande schildklierproblemen kan extra jodium klachten verergeren."},
    {"id":"CF006","keywords":["jodium","kelp","blaaswier","fucus vesiculosus"],"operator":"gte","threshold":500,"unit":"mcg","severity":"Critical","message":"Zeer hoge jodium-inname. Boven 500 mcg per dag bestaat risico op schildklierproblematiek, ook zonder voorgeschiedenis."},
    {"id":"CF007","keywords":["ashwagandha","withania somnifera"],"operator":"any","threshold":0,"unit":"","severity":"Major","message":"Ashwagandha kan de schildklierfunctie beïnvloeden. Bij schildklierproblematiek of schildkliermedicatie: gebruik alleen onder begeleiding."},
    {"id":"CF008","keywords":["ijzer","ferrum","feso4","ferrousfumaraat","ijzerbisglycinaat","iron bisglycinate","ferrous"],"operator":"any","threshold":0,"unit":"","severity":"Major","message":"IJzersuppletie is alleen zinvol bij aangetoond tekort. Overmatige ijzerinname is geassocieerd met oxidatieve stress. Niet aanbevolen zonder indicatie."},
    {"id":"CF009","keywords":["ijzer","ferrum","iron","ferrous"],"operator":"gte","threshold":45,"unit":"mg","severity":"Critical","message":"Hoge ijzerdosering. Boven 45 mg per dag verhoogd risico op bijwerkingen. Alleen onder medische supervisie."},
    {"id":"CF010","keywords":["zink","zinc","zinkbisglycinaat","zinc bisglycinate"],"operator":"gte","threshold":25,"unit":"mg","severity":"Major","message":"Hoge zink-inname kan koper verdringen. Langdurig gebruik zonder koper-suppletie kan leiden tot koperdepletie."},
    {"id":"CF011","keywords":["zink","zinc"],"operator":"gte","threshold":40,"unit":"mg","severity":"Critical","message":"Zinkdosering nadert de veiligheidsgrens van 40 mg per dag. Niet combineren met andere zinkbevattende supplementen."},
    {"id":"CF012","keywords":["calcium","calciumcarbonaat","calciumcitraat","calcium carbonate","calcium citrate"],"operator":"gte","threshold":1200,"unit":"mg","severity":"Major","message":"Hoge calcium-inname kan de opname van ijzer, zink en magnesium verminderen. Neem niet gelijktijdig in met andere mineralen."},
    {"id":"CF013","keywords":["selenium","natriumseleniet","selenomethionine","sodium selenite"],"operator":"gte","threshold":200,"unit":"mcg","severity":"Major","message":"Selenium heeft een smal therapeutisch venster. Chronisch hoge inname is geassocieerd met selenose."},
    {"id":"CF014","keywords":["vitamine k","vitamin k","k2","mk-7","mk7","menaquinon","menaquinone","phylloquinon","phylloquinone"],"operator":"any","threshold":0,"unit":"","severity":"Critical","message":"Vitamine K interfereert direct met bloedverdunners zoals Acenocoumarol en Warfarine. Bij gebruik van bloedverdunners: overleg altijd met je arts."},
    {"id":"CF015","keywords":["epa","dha","omega-3","omega 3","visolie","fish oil","levertraan","krillolie","krill oil"],"operator":"gte","threshold":3000,"unit":"mg","severity":"Major","message":"Hoge omega-3 dosering heeft een mild bloedverdunnend effect. Bij antistollingsmedicatie of geplande operatie: overleg met je arts."},
    {"id":"CF016","keywords":["knoflookextract","allium sativum","ginkgo biloba","ginkgo extract"],"operator":"any","threshold":0,"unit":"","severity":"Major","message":"Dit ingrediënt heeft bloedverdunnende eigenschappen. Niet combineren met antistollingsmedicatie zonder medisch advies."},
    {"id":"CF017","keywords":["sint-janskruid","hypericum perforatum","st. john","st john"],"operator":"any","threshold":0,"unit":"","severity":"Critical","message":"Sint-Janskruid verlaagt de werkzaamheid van veel medicijnen waaronder anticonceptie, antidepressiva en bloedverdunners. Niet combineren met medicijnen."},
    {"id":"CF018","keywords":["curcumine","curcuma longa","kurkuma extract"],"operator":"gte","threshold":500,"unit":"mg","severity":"Major","message":"Hoge curcumine-dosering heeft anticoagulerend effect. Bij galstenen of antistollingsmedicatie: gebruik voorzichtig."},
    {"id":"CF019","keywords":["retinol","vitamine a"],"operator":"gte","threshold":800,"unit":"mcg","severity":"Critical","message":"Vitamine A boven aanbevolen zwangerschapsdosering. Hoog risico bij zwangerschap."},
    {"id":"CF020","keywords":["ashwagandha","withania somnifera","rhodiola","rhodiola rosea"],"operator":"any","threshold":0,"unit":"","severity":"Critical","message":"Adaptogenen zijn onvoldoende onderzocht op veiligheid tijdens zwangerschap en borstvoeding. Gebruik wordt afgeraden."},
    {"id":"CF021","keywords":["vitamine b6","pyridoxine","p-5-p"],"operator":"gte","threshold":10,"unit":"mg","severity":"Major","message":"B6 boven 10 mg per dag tijdens zwangerschap: afstemmen met arts."},
    {"id":"CF022","keywords":["foliumzuur","folaat","5-mthf","methylfolaat","folic acid","methylfolate"],"operator":"any","threshold":0,"unit":"","severity":"Info","message":"Dit product bevat folaat. Relevant bij zwangerschapswens en zwangerschap. Controleer of de dosering minimaal 400 mcg per dag is."},
    {"id":"CF023","keywords":["vitamine d3","cholecalciferol","vitamin d3"],"operator":"gte","threshold":4000,"unit":"IU","severity":"Major","message":"Hoge vitamine D3 dosering. Bij zwangerschap of nierfunctiestoornissen: gebruik niet zonder medisch advies."},
    {"id":"CF024","keywords":["vitamine c","ascorbinezuur","natriumascorbaat","vitamin c","ascorbic acid"],"operator":"gte","threshold":1000,"unit":"mg","severity":"Major","message":"Hoge vitamine C dosering verhoogt de oxalaatuitscheiding via de nier. Bij nierstenen of verminderde nierfunctie: gebruik voorzichtig."},
    {"id":"CF025","keywords":["magnesiumoxide","magnesiumcitraat","magnesiumbisglycinaat","magnesium bisglycinate","magnesium malate","magnesium threonate"],"operator":"gte","threshold":350,"unit":"mg","severity":"Major","message":"Hoge magnesium-inname. Bij nierfunctiestoornissen kan magnesium niet goed worden uitgescheiden."},
    {"id":"CF026","keywords":["creatine monohydraat","creatine monohydrate","creatine"],"operator":"any","threshold":0,"unit":"","severity":"Minor","message":"Creatine verhoogt creatinine-waarden in het bloed wat bloedtesten kan vertekenen. Bij verminderde nierfunctie: gebruik alleen na medisch overleg."},
    {"id":"CF027","keywords":["groene thee extract","egcg","camellia sinensis extract"],"operator":"gte","threshold":400,"unit":"mg","severity":"Major","message":"Geconcentreerde groene thee-extracten zijn in zeldzame gevallen geassocieerd met leverschade bij nuchter gebruik."},
    {"id":"CF028","keywords":["nicotinezuur","nicotinic acid","niacine flush"],"operator":"gte","threshold":500,"unit":"mg","severity":"Major","message":"Hoge niacine-dosering kan flush veroorzaken en bij langdurig gebruik lever-enzymstoringen geven."},
    {"id":"CF029","keywords":["informed sport","nzvt","wada gecertificeerd","doping gecertificeerd"],"operator":"any","threshold":0,"unit":"","severity":"Critical","message":"Dit product heeft geen aantoonbare doping-certificering. Voor competitieve sporters bestaat risico op een onbedoelde positieve dopingtest."},
    {"id":"CF030","keywords":["creatine monohydraat","creatine monohydrate","creapure"],"operator":"any","threshold":0,"unit":"","severity":"Minor","message":"Creatine is een toegestaan supplement voor sporters. Controleer altijd of het specifieke product een doping-keurmerk heeft."},
    {"id":"CF031","keywords":["cafeine","caffeine","cafeïne anhydraat","caffeine anhydrous"],"operator":"gte","threshold":200,"unit":"mg","severity":"Major","message":"Hoge cafeïne-dosering. Bij hart- en vaataandoeningen, slaapproblemen of angstklachten: gebruik voorzichtig."},
    {"id":"CF032","keywords":["beta-alanine"],"operator":"gte","threshold":3200,"unit":"mg","severity":"Minor","message":"Beta-alanine kan tintelingen veroorzaken bij hogere doseringen. Niet gevaarlijk maar kan als oncomfortabel worden ervaren."},
    {"id":"CF033","keywords":["berberine","berberis vulgaris extract"],"operator":"any","threshold":0,"unit":"","severity":"Critical","message":"Berberine heeft een significant bloedsuikerverlagend effect. Combinatie met diabetesmedicatie kan een gevaarlijk lage bloedsuiker veroorzaken."},
    {"id":"CF034","keywords":["chroompicolinaat","chromium picolinate","chroom","chromium"],"operator":"gte","threshold":200,"unit":"mcg","severity":"Major","message":"Chroom kan insulinegevoeligheid verbeteren. Bij gelijktijdig gebruik van insuline of diabetesmedicatie: risico op hypoglykemie."},
    {"id":"CF035","keywords":["alfa-liponzuur","alpha lipoic acid","r-liponzuur","r-ala"],"operator":"gte","threshold":300,"unit":"mg","severity":"Major","message":"Alfa-liponzuur verlaagt bloedsuiker. Bij gebruik van diabetesmedicatie of schildkliermedicatie: controleer dosering."},
    {"id":"CF036","keywords":["foliumzuur","folic acid","pteroylglutaminezuur"],"operator":"any","threshold":0,"unit":"","severity":"Major","message":"Synthetisch foliumzuur kan bij MTHFR-polymorfisme onvoldoende worden omgezet. De actieve vorm methylfolaat is de betere keuze."},
    {"id":"CF037","keywords":["cyanocobalamine","cyanocobalamin"],"operator":"any","threshold":0,"unit":"","severity":"Minor","message":"Cyanocobalamine vereist een extra omzettingsstap. Bij MTHFR-polymorfisme is methylcobalamine de voorkeursvorm."},
    {"id":"CF038","keywords":["s-adenosyl-l-methionine","same supplement","s-adenosylmethionine"],"operator":"any","threshold":0,"unit":"","severity":"Major","message":"SAMe is een methyldonor en kan serotoninespiegel verhogen. Bij antidepressiva of bipolaire stoornis: niet zonder begeleiding gebruiken."},
    {"id":"CF039","keywords":["echinacea purpurea","echinacea angustifolia","echinacea extract"],"operator":"any","threshold":0,"unit":"","severity":"Major","message":"Echinacea stimuleert het immuunsysteem. Bij auto-immuunziekten of immunosuppressiva: gebruik wordt afgeraden."},
    {"id":"CF040","keywords":["astragalus membranaceus","astragalus extract","huang qi"],"operator":"any","threshold":0,"unit":"","severity":"Major","message":"Astragalus is een immuunstimulant. Bij auto-immuunziekten of immunosuppressiva: vermijd dit supplement."},
    {"id":"CF041","keywords":["vitamine d3","cholecalciferol","vitamin d3"],"operator":"gte","threshold":2000,"unit":"IU","severity":"Info","message":"Vitamine D speelt een rol bij immuunregulatie. Bij auto-immuunziekten kan hogere vitamine D therapeutisch relevant zijn. Bespreek dit met je arts."},
    {"id":"CF042","keywords":["co-enzym q10","coenzyme q10","ubiquinol","ubiquinone","coq10"],"operator":"any","threshold":0,"unit":"","severity":"Major","message":"CoQ10 is extra relevant bij statinegebruik. Statines verlagen de aanmaak van CoQ10 in het lichaam."},
    {"id":"CF043","keywords":["rode gistrijst extract","monascus purpureus","monacoline k","red yeast rice"],"operator":"any","threshold":0,"unit":"","severity":"Critical","message":"Rode gistrijst bevat een natuurlijk statine. Combinatie met statinetherapie geeft verhoogd risico op spierschade en leverbelasting."},
    {"id":"CF044","keywords":["magnesiumbisglycinaat","magnesium bisglycinate","magnesiumcitraat","magnesium citrate","magnesiummalaat"],"operator":"any","threshold":0,"unit":"","severity":"Info","message":"Magnesium heeft een gunstige rol bij bloeddrukregulatie en hartritme. Relevant bij hypertensie of hartritmeproblematiek."},
    {"id":"CF045","keywords":["melatonine","melatonin"],"operator":"gte","threshold":0.5,"unit":"mg","severity":"Major","message":"Melatonine kan interacteren met antidepressiva en slaap- of kalmeringsmiddelen. Bij gebruik van deze medicijnen: raadpleeg je arts."},
    {"id":"CF046","keywords":["5-htp","5-hydroxytryptofaan","5-hydroxytryptophan"],"operator":"any","threshold":0,"unit":"","severity":"Critical","message":"5-HTP verhoogt de serotoninesynthese. Niet combineren met antidepressiva of MAO-remmers: risico op serotoninesyndroom."},
    {"id":"CF047","keywords":["l-tryptofaan","l-tryptophan","tryptophan"],"operator":"gte","threshold":500,"unit":"mg","severity":"Critical","message":"Hoge tryptofaan-dosering verhoogt serotonineproductie. Bij gebruik van SSRI of MAO-remmers: risico op serotoninesyndroom."},
    {"id":"CF048","keywords":["gamma-aminoboterzuur","gaba supplement","gamma aminobutyric acid"],"operator":"any","threshold":0,"unit":"","severity":"Minor","message":"GABA-supplementen kunnen het centrale zenuwstelsel beïnvloeden. Bij benzodiazepinen of gabapentinoïden: gebruik voorzichtig."},
    {"id":"CF049","keywords":["valeriaan extract","valeriana officinalis","valerian root"],"operator":"any","threshold":0,"unit":"","severity":"Minor","message":"Valeriaan heeft milde sedatieve eigenschappen. Bij slaap- of kalmeringsmiddelen: additief effect mogelijk."},
    {"id":"CF050","keywords":["lactobacillus","bifidobacterium","probiotisch","probiotic blend","probiotica"],"operator":"any","threshold":0,"unit":"","severity":"Critical","message":"Levende bacteriën in probiotica kunnen bij ernstig immuungecompromitteerde personen systemische infecties veroorzaken. Raadpleeg je arts bij chemotherapie of transplantatie."},
    {"id":"CF051","keywords":["saccharomyces boulardii","saccharomyces cerevisiae"],"operator":"any","threshold":0,"unit":"","severity":"Major","message":"Saccharomyces is een gist. Bij gebruik van antifunginale middelen worden gistprobiotica geïnactiveerd en is suppletie ineffectief."},
    {"id":"CF052","keywords":["voor kinderen","children","pediatric","kids formula"],"operator":"any","threshold":0,"unit":"","severity":"Critical","message":"Dit product is niet specifiek geformuleerd of gedoseerd voor kinderen. Doseringen voor volwassenen zijn niet direct toepasbaar voor kinderen."},
    {"id":"CF053","keywords":["vitamine d3","cholecalciferol","vitamin d3"],"operator":"gte","threshold":1000,"unit":"IU","severity":"Major","message":"Vitamine D dosering boven 1000 IU per dag voor kinderen: controleer of dit past bij de leeftijdsspecifieke behoefte."},
    {"id":"CF054","keywords":["vitamine k2","mk-7","mk7","menaquinon-7","menaquinone-7"],"operator":"any","threshold":0,"unit":"","severity":"Info","message":"Vitamine K2 is relevant voor botgezondheid en cardiovasculaire bescherming bij senioren. Controleer interactie met eventuele antistollingsmedicatie."},
    {"id":"CF055","keywords":["vitamine e","mixed tocopherols","d-alpha tocopherol","dl-alpha tocopherol"],"operator":"gte","threshold":400,"unit":"IU","severity":"Major","message":"Hoge vitamine E heeft een antistollend effect. Bij bloedverdunners is er een verhoogd risico op bloedingen."},
    {"id":"CF056","keywords":["beta-caroteen","beta-carotene","provitamine a"],"operator":"gte","threshold":5000,"unit":"IU","severity":"Critical","message":"Hoge beta-caroteen dosering bij rokers is geassocieerd met verhoogd longkankerrisico. Hoge retinol bij zwangerschap heeft teratogeen risico."},
    {"id":"CF057","keywords":["vitamine d3","cholecalciferol","vitamin d3"],"operator":"gte","threshold":4000,"unit":"IU","severity":"Major","message":"Zeer hoge vitamine D3 dosering. Vetoplosbare vitaminen stapelen bij langdurig hoge inname. Controleer serumspiegels bij gebruik boven 4000 IU per dag."},
    {"id":"CF058","keywords":["vitamine d3","cholecalciferol","vitamin d","vitamine d"],"operator":"any","threshold":0,"unit":"","severity":"Info","message":"In de Nederlandse winter is zonlicht onvoldoende voor vitamine D-aanmaak. Een dosering onder 1000 IU per dag kan onvoldoende zijn zonder zonblootstelling."},
    {"id":"CF059","keywords":["vitamine d3","cholecalciferol","vitamin d3"],"operator":"gte","threshold":2000,"unit":"IU","severity":"Info","message":"In de zomer met voldoende buitenactiviteit kan endogene vitamine D aanmaak voldoende zijn. Hoge suppletie gecombineerd met zonblootstelling kan bij gevoelige personen tot hypervitaminose leiden."},
]

EXCIPIENT_KEYWORDS = [
    "capsulehuls", "hydroxypropylmethylcellulose", "hpmc", "rijstvezel",
    "antiklontermiddel", "bevochtigingsmiddel", "glycerol", "gezuiverd water",
    "siliciumdioxide", "magnesiumstearaat", "talk", "titaniumdioxide",
    "gelatine", "visgelatine", "zetmeel", "maltodextrine", "silica",
    "cellulose", "stearinezuur", "carrageen", "lecithine", "arabische gom",
    "dicalciumfosfaat", "microkristallijne", "magnesium stearate",
    "stearic acid", "silicon dioxide", "microcrystalline cellulose",
    "hydroxypropyl methylcellulose", "gelatin", "dicalcium phosphate",
    "calcium carbonate as filler", "rice flour", "rice bran", "talc",
    "titanium dioxide", "croscarmellose", "povidone", "polyethylene glycol",
    "polysorbate", "sodium lauryl sulfate", "maltodextrin", "olijfolie",
    "olive oil", "sunflower oil", "zonnebloemolie", "beeswax", "bijenwas",
    "carnauba wax", "carnauba was", "shellac", "schellak",
]

CERT_KEYWORDS = [
    "IFOS", "ISO 17025", "ISO 22000", "GMP", "HACCP", "USP", "NSF",
    "Informed Sport", "NZVT", "Creapure", "MSC", "Rainforest Alliance",
    "Fair Trade", "Friend of the Sea", "Dolphin Safe", "Green-e",
    "TRAACS", "Albion", "Golden Omega", "WADA", "Kosher", "Halal",
    "Vegan Society", "Soil Association", "ISO 9001", "FSSC 22000",
    "BRC", "IFS", "organic", "biologisch", "Keurmerk",
]

JARGON_MAP = [
    (r"\bbiobeschikbaarheid\b", "opneembaarheid"),
    (r"\btherapeutisch\b", "werkzaam"),
    (r"\bklinisch relevante\b", "werkzame"),
    (r"\bklinisch bewezen\b", ""),
    (r"\bstudies tonen aan\b", ""),
    (r"\bfarmaceutisch\b", "van hoge kwaliteit"),
    (r"\bchelaatvorm\b", "goed opneembare gebonden vorm"),
    (r"\bCOA\b", "onafhankelijk laboratoriumrapport"),
    (r"\bTOTOX\b", "versheidswaarde van de olie"),
    (r"\bproprietary blend\b", "mengsel met verborgen doseringen"),
    (r"\binflammatie\b", "ontsteking"),
    (r"\bsystemisch\b", "via het bloed"),
    (r"\bendogeen\b", "eigen aanmaak"),
    (r"\bexogeen\b", "via suppletie"),
]

AMOUNT_PAT = re.compile(
    r"(\d[\d.,]*)\s*(mg|g|mcg|µg|ug|ml|ie|iu|kve|cfu|%)", re.I
)

# ═══════════════════════════════════════════════
# DEEL 3 — EXCEL LOADING
# ═══════════════════════════════════════════════

def load_all_data():
    """Laad alle Excel data eenmalig bij startup."""

    # Engine Review V4 — criteria en CF rules
    engine_path = os.path.join(BASE_DIR, "Funcify. Engine Review V4.xlsx")
    if os.path.exists(engine_path) and HAS_OPENPYXL:
        try:
            wb = openpyxl.load_workbook(engine_path, data_only=True)

            criteria = []
            if "Engine_review_score" in wb.sheetnames:
                ws = wb["Engine_review_score"]
                header_row = None
                for row in ws.iter_rows():
                    vals = [str(c.value or "").strip() for c in row]
                    if "Category" in vals and "Criterion" in vals and "Weight" in vals:
                        header_row = {v: i for i, v in enumerate(vals)}
                        break
                if header_row:
                    for row in ws.iter_rows():
                        vals = [c.value for c in row]
                        category = str(vals[header_row.get("Category", 0)] or "")
                        if category in ["Category", "SUMMARY", "INPUT",
                                        "HELPERS", "MODEL SETTINGS", ""]:
                            continue
                        criterion = str(vals[header_row.get("Criterion", 1)] or "")
                        weight_raw = vals[header_row.get("Weight", 2)]
                        critical = str(vals[header_row.get("Critical? (Y/N)", 3)] or "N").strip().upper()
                        applies_to = str(vals[header_row.get("AppliesTo", 4)] or "ALL").strip()
                        try:
                            weight = float(weight_raw) if weight_raw else 0
                        except (ValueError, TypeError):
                            weight = 0
                        if criterion and weight > 0:
                            criteria.append({
                                "source": "core",
                                "category": category,
                                "criterion": criterion,
                                "weight": weight,
                                "critical": critical == "Y",
                                "applies_to": applies_to,
                            })

            if "Category_modules" in wb.sheetnames:
                ws = wb["Category_modules"]
                header_row = None
                for row in ws.iter_rows():
                    vals = [str(c.value or "").strip() for c in row]
                    if "ModuleType" in vals and "Criterion" in vals and "Weight" in vals:
                        header_row = {v: i for i, v in enumerate(vals)}
                        break
                if header_row:
                    for row in ws.iter_rows():
                        vals = [c.value for c in row]
                        module_type = str(vals[header_row.get("ModuleType", 0)] or "").strip()
                        if not module_type or module_type in ["ModuleType", "Producttype", ""]:
                            continue
                        criterion = str(vals[header_row.get("Criterion", 2)] or "")
                        weight_raw = vals[header_row.get("Weight", 3)]
                        critical = str(vals[header_row.get("Critical", 4)] or "N").strip().upper()
                        try:
                            weight = float(weight_raw) if weight_raw else 0
                        except (ValueError, TypeError):
                            weight = 0
                        if criterion and weight > 0:
                            criteria.append({
                                "source": "module",
                                "category": f"Module_{module_type}",
                                "criterion": criterion,
                                "weight": weight,
                                "critical": critical == "Y",
                                "applies_to": module_type,
                            })

            _cache["criteria"] = criteria
            print(f"ENGINE CRITERIA GELADEN: {len(criteria)}", flush=True)
        except Exception as e:
            print(f"ENGINE EXCEL ERROR: {e}", flush=True)
            _cache["criteria"] = []
    else:
        print("ENGINE EXCEL NIET GEVONDEN", flush=True)
        _cache["criteria"] = []

    # Consumer UI — ui_display_content laden
    ui_path = os.path.join(BASE_DIR, "Funcify. Consumer UI.xlsx")
    if os.path.exists(ui_path) and HAS_OPENPYXL:
        try:
            wb = openpyxl.load_workbook(ui_path, data_only=True)
            ui_content = {}
            if "ui_display_content" in wb.sheetnames:
                ws = wb["ui_display_content"]
                headers = None
                for row in ws.iter_rows():
                    vals = [str(c.value or "").strip() for c in row]
                    if "ingredient_id" in vals:
                        headers = {v: i for i, v in enumerate(vals)}
                        continue
                    if headers and vals[0]:
                        ing_id = vals[headers.get("ingredient_id", 0)]
                        if ing_id:
                            ui_content[ing_id] = {
                                "display_name": vals[headers.get("display_name", 1)] if headers.get("display_name") is not None else "",
                                "short_benefit": vals[headers.get("short_benefit", 2)] if headers.get("short_benefit") is not None else "",
                                "why_it_helps": vals[headers.get("why_it_helps", 3)] if headers.get("why_it_helps") is not None else "",
                                "when_to_use": vals[headers.get("when_to_use", 4)] if headers.get("when_to_use") is not None else "",
                                "caution_text": vals[headers.get("caution_text", 5)] if headers.get("caution_text") is not None else "",
                                "preferred_format_text": vals[headers.get("preferred_format_text", 6)] if headers.get("preferred_format_text") is not None else "",
                            }
            _cache["ui_content"] = ui_content
            print(f"UI CONTENT GELADEN: {len(ui_content)} ingrediënten", flush=True)
        except Exception as e:
            print(f"UI EXCEL ERROR: {e}", flush=True)
            _cache["ui_content"] = {}
    else:
        _cache["ui_content"] = {}


with app.app_context():
    load_all_data()

# ═══════════════════════════════════════════════
# DEEL 4 — HELPER FUNCTIES
# ═══════════════════════════════════════════════

def simplify_jargon(obj):
    if isinstance(obj, str):
        for pattern, replacement in JARGON_MAP:
            obj = re.sub(pattern, replacement, obj, flags=re.I)
        return obj
    elif isinstance(obj, list):
        return [simplify_jargon(i) for i in obj]
    elif isinstance(obj, dict):
        return {k: simplify_jargon(v) for k, v in obj.items()}
    return obj


def _unit_to_base(value, unit):
    unit = unit.lower().strip()
    if unit in ("mcg", "µg", "ug"):
        return float(value), "mcg"
    if unit in ("iu", "ie"):
        return float(value), "IU"
    if unit == "g":
        return float(value) * 1000, "mg"
    return float(value), unit


def _parse_price(price_str):
    if not price_str:
        return 0.0
    cleaned = re.sub(r"[^\d,\.]", "", str(price_str)).replace(",", ".")
    m = re.search(r"\d+\.\d+|\d+", cleaned)
    return float(m.group(0)) if m else 0.0


def _parse_units(package_size_str):
    m = re.search(r"(\d+)", str(package_size_str or ""))
    return int(m.group(1)) if m else 0


def _parse_servings_per_day(serving_size_str, usage_instructions_str=""):
    s = str(serving_size_str or "")
    u = str(usage_instructions_str or "")
    m = re.search(r"(\d+)", s)
    spd = max(1, int(m.group(1))) if m else 1
    if spd == 1 and re.search(r"\b2\b.{0,20}(per dag|dagelijks)", u, re.I):
        return 2
    return min(spd, 10)


def calculate_price_per_day(price_str, package_size_str,
                             serving_size_str, usage_str=""):
    price = _parse_price(price_str)
    units = _parse_units(package_size_str)
    spd = _parse_servings_per_day(serving_size_str, usage_str)
    if not price or not units:
        return ""
    days = units / spd
    if days <= 0:
        return ""
    return "€" + f"{price/days:.2f}".replace(".", ",") + " per dag"


def calculate_price_per_gram(price_str, ingredients,
                              package_size_str, serving_size_str):
    price = _parse_price(price_str)
    units = _parse_units(package_size_str)
    spd = _parse_servings_per_day(serving_size_str)
    if not price or not units:
        return ""
    total_mg = sum(
        float(i.get("amount") or 0) * (1000 if str(i.get("unit", "")).lower() == "g" else 1)
        for i in ingredients
        if i.get("type") != "vul-additief" and i.get("amount") is not None
        and str(i.get("unit", "")).lower() in ("mg", "g")
    )
    if total_mg <= 0:
        return ""
    total_g = (total_mg * units) / spd / 1000
    if total_g <= 0:
        return ""
    return "€" + f"{price/total_g:.2f}".replace(".", ",") + " per gram"


def get_intake_advice(product_data):
    """Alleen van het label. Nooit fabriceren."""
    usage = str(product_data.get("usage_instructions") or "").strip()
    if usage and len(usage) > 10:
        return usage
    return ""


def _split_active_excipients(ingredients):
    active = [i for i in ingredients if i.get("type") != "vul-additief"]
    excipients = [i for i in ingredients if i.get("type") == "vul-additief"]
    return active, excipients

# ═══════════════════════════════════════════════
# DEEL 5 — CONTEXT FLAGS EN COFACTOREN
# ═══════════════════════════════════════════════

def evaluate_context_flags(ingredients):
    """
    Evalueert CF-rules ALLEEN op basis van ingrediënten die
    daadwerkelijk in het product zitten.
    Geeft dicts terug met id, severity, message.
    Gesorteerd op severity.
    """
    triggered = []
    triggered_ids = set()

    for rule in CONTEXT_FLAG_RULES:
        rule_id = rule["id"]
        if rule_id in triggered_ids:
            continue

        for ing in ingredients:
            if ing.get("type") == "vul-additief":
                continue

            name_lower = str(ing.get("name", "")).lower()
            matched = any(kw in name_lower for kw in rule["keywords"])

            if not matched:
                continue

            if rule["operator"] == "any":
                triggered.append({
                    "id": rule_id,
                    "severity": rule["severity"],
                    "message": rule["message"]
                })
                triggered_ids.add(rule_id)
                break

            elif rule["operator"] == "gte":
                amount = ing.get("amount")
                unit = str(ing.get("unit", ""))
                if amount is None or not unit:
                    continue
                try:
                    val, norm_unit = _unit_to_base(float(amount), unit)
                    _, norm_thresh = _unit_to_base(rule["threshold"], rule["unit"])
                    if norm_unit == norm_thresh and val >= rule["threshold"]:
                        triggered.append({
                            "id": rule_id,
                            "severity": rule["severity"],
                            "message": rule["message"]
                        })
                        triggered_ids.add(rule_id)
                        break
                except (ValueError, TypeError):
                    continue

    triggered.sort(key=lambda f: SEVERITY_ORDER.get(f.get("severity", "Info"), 4))
    return triggered


def evaluate_cofactor_checks(ingredients):
    """
    Controleert kritische cofactorparen.
    Geeft dicts terug met id, severity, message.
    """
    warnings = []
    names_lower = [
        str(i.get("name", "")).lower()
        for i in ingredients
        if i.get("type") != "vul-additief"
    ]

    def has_ing(keywords):
        return any(any(kw in n for kw in keywords) for n in names_lower)

    def get_amount_unit(keywords, target_unit):
        for ing in ingredients:
            if ing.get("type") == "vul-additief":
                continue
            n = str(ing.get("name", "")).lower()
            if any(kw in n for kw in keywords):
                amt = ing.get("amount")
                u = str(ing.get("unit") or "").lower()
                if amt is not None:
                    try:
                        val, norm = _unit_to_base(float(amt), u)
                        _, norm_target = _unit_to_base(0, target_unit)
                        if norm == norm_target:
                            return val
                    except (ValueError, TypeError):
                        pass
        return 0.0

    # D3 zonder K2
    d3_iu = get_amount_unit(
        ["vitamine d3", "cholecalciferol", "vitamin d3"], "IU"
    )
    k2_present = has_ing(["vitamine k2", "vitamin k2", "k2", "mk-7",
                           "mk7", "menaquinon", "menaquinone"])
    if d3_iu >= 1000 and not k2_present:
        warnings.append({
            "id": "COF001", "severity": "Major",
            "message": "Hoge vitamine D3 zonder K2. K2 zorgt dat calcium op de juiste plek terechtkomt. Overweeg een product met K2 erbij."
        })

    # Hoog zink zonder koper
    zinc_mg = get_amount_unit(
        ["zink", "zinc", "zinkbisglycinaat", "zinc bisglycinate"], "mg"
    )
    copper_present = has_ing(["koper", "copper", "cuprum",
                               "koperbisglycinaat", "copper bisglycinate"])
    if zinc_mg >= 25 and not copper_present:
        warnings.append({
            "id": "COF002", "severity": "Major",
            "message": "Hoog zink zonder koper kan bij langdurig gebruik koperdepletie veroorzaken wat de bloedvorming beïnvloedt."
        })

    # IJzer zonder vitamine C
    iron_present = has_ing(["ijzer", "iron", "ferrum", "ferrous",
                             "ijzerbisglycinaat", "iron bisglycinate"])
    vitc_present = has_ing(["vitamine c", "ascorbinezuur",
                             "vitamin c", "ascorbic acid"])
    if iron_present and not vitc_present:
        warnings.append({
            "id": "COF003", "severity": "Minor",
            "message": "Vitamine C verbetert ijzeropname significant. Combineer bij voorkeur met vitamine C of neem in met vitamine C-rijke voeding."
        })

    # Omega-3 zonder vitamine E
    omega3_present = has_ing(["epa", "dha", "omega-3", "omega 3",
                               "visolie", "fish oil"])
    vite_present = has_ing(["vitamine e", "tocoferol", "tocopherol",
                             "vitamin e", "mixed tocopherols"])
    if omega3_present and not vite_present:
        warnings.append({
            "id": "COF004", "severity": "Major",
            "message": "Omega-3 vetzuren zijn gevoelig voor oxidatie. Controleer of het product antioxidantbescherming bevat zoals vitamine E."
        })

    # Calcium en ijzer samen — antagonisme
    calcium_present = has_ing(["calcium", "calciumcarbonaat", "calciumcitraat"])
    if calcium_present and iron_present:
        warnings.append({
            "id": "ANT001", "severity": "Major",
            "message": "Calcium en ijzer in hetzelfde product verminderen elkaars opname. Neem ze op verschillende momenten in voor optimaal effect."
        })

    warnings.sort(key=lambda f: SEVERITY_ORDER.get(f.get("severity", "Info"), 4))
    return warnings

# ═══════════════════════════════════════════════
# DEEL 6 — SCRAPER
# ═══════════════════════════════════════════════

def fetch_html(url):
    response = requests.get(url, headers=HEADERS, timeout=20, allow_redirects=True)
    response.raise_for_status()
    return BeautifulSoup(response.content, "lxml"), response.text


def fetch_with_scrapingbee(url):
    """
    Drie niveaus:
    1. ScrapingBee standaard met JS rendering
    2. ScrapingBee premium proxy als standaard faalt
    3. Directe fetch als fallback
    Geeft tuple terug: (BeautifulSoup, raw_html_text)
    """
    api_key = os.environ.get("SCRAPINGBEE_KEY")

    if not api_key:
        print("SCRAPER: geen ScrapingBee key, directe fetch", flush=True)
        return fetch_html(url)

    # Niveau 1: standaard ScrapingBee
    try:
        response = requests.get(
            "https://app.scrapingbee.com/api/v1/",
            params={
                "api_key": api_key,
                "url": url,
                "render_js": "true",
                "wait": "4000",
                "block_ads": "true",
                "block_resources": "true",
                "country_code": "nl",
            },
            timeout=45,
        )
        response.raise_for_status()
        content = response.content
        text = response.text
        print(f"SCRAPER: ScrapingBee standaard geslaagd ({len(text)} tekens)", flush=True)
        return BeautifulSoup(content, "lxml"), text
    except Exception as e:
        print(f"SCRAPER: ScrapingBee standaard gefaald ({e}), probeer premium", flush=True)

    # Niveau 2: premium proxy
    try:
        response = requests.get(
            "https://app.scrapingbee.com/api/v1/",
            params={
                "api_key": api_key,
                "url": url,
                "render_js": "true",
                "wait": "4000",
                "block_ads": "true",
                "block_resources": "true",
                "premium_proxy": "true",
                "country_code": "nl",
            },
            timeout=45,
        )
        response.raise_for_status()
        content = response.content
        text = response.text
        print(f"SCRAPER: ScrapingBee premium geslaagd ({len(text)} tekens)", flush=True)
        return BeautifulSoup(content, "lxml"), text
    except Exception as e:
        print(f"SCRAPER: ScrapingBee premium gefaald ({e}), directe fetch", flush=True)

    # Niveau 3: directe fetch
    return fetch_html(url)


def scrape_product(soup, url=""):
    """
    Extraheert basisdata via HTML selectors.
    Raadt NOOIT iets. Laat velden leeg als niet gevonden.
    """
    page_text = soup.get_text(" ", strip=True)

    # Productnaam
    product_name = ""
    for sel in ["h1", ".product-title", ".product-name",
                '[itemprop="name"]', ".pdp-title"]:
        els = soup.select(sel)
        if els:
            candidate = els[0].get_text(strip=True)[:200]
            if candidate:
                product_name = candidate
                break

    # Merk
    brand_name = ""
    for sel in ['[itemprop="brand"]', '[itemprop="manufacturer"]',
                ".brand", ".manufacturer", '[class*="brand"]',
                '[class*="vendor"]']:
        els = soup.select(sel)
        if els:
            candidate = els[0].get_text(strip=True)[:100]
            if candidate and len(candidate) < 60:
                brand_name = candidate
                break
    if not brand_name:
        for meta in soup.find_all("meta"):
            prop = meta.get("property", "") + meta.get("name", "")
            if "brand" in prop.lower() or "og:site_name" in prop.lower():
                content = meta.get("content", "").strip()
                if content and len(content) < 60:
                    brand_name = content
                    break

    # Prijs
    price = ""
    price_match = re.search(r"[€\$]?\s*(\d+[,\.]\d{2})", page_text)
    if price_match:
        price = f"€{price_match.group(1)}"

    # Package size — alleen uit productnaam of expliciete vermelding
    package_size = ""
    unit_pat = r"(\d+)\s*(softgels?|capsules?|tabletten?|vegicaps?|caps?|stuks?)"
    m = re.search(unit_pat, product_name, re.I)
    if m:
        package_size = m.group(0)

    # Certificeringen via tekst matching
    certifications = []
    text_lower = page_text.lower()
    for kw in CERT_KEYWORDS:
        if kw.lower() in text_lower and kw not in certifications:
            certifications.append(kw)

    return {
        "product_name": product_name,
        "brand_name": brand_name,
        "price": price,
        "package_size": package_size,
        "certifications": certifications,
        "ingredients": [],
        "active_ingredients": [],
        "excipients": [],
        "serving_size": "",
        "usage_instructions": "",
        "health_claims": [],
        "warnings": [],
        "additional_info": page_text[:8000],
    }


def extract_with_claude(url, raw_text, product_data):
    """
    Claude leest de volledige paginatekst en extraheert
    gestructureerde data.
    Regels:
    - Alleen invullen als letterlijk op de pagina staat
    - source_text verplicht per ingredient
    - Nooit gokken of aannames maken
    """
    client = anthropic.Anthropic(api_key=os.environ.get("ANTHROPIC_API_KEY"))

    full_text = raw_text[:15000] if raw_text else ""

    # Debug: zoek waar ingrediënten staan in de volledige tekst
    for keyword in ["cholecalciferol", "menaquinon", "EPA", "DHA", "vitamine d3"]:
        pos = raw_text.lower().find(keyword.lower()) if raw_text else -1
        if pos >= 0:
            print(f"DEBUG: '{keyword}' gevonden op positie {pos} van {len(raw_text)}", flush=True)

    system_prompt = """Je extraheert productinformatie van supplement-paginas
voor de Funcify beoordelingsengine.

ABSOLUTE REGELS:
1. Vul een veld ALLEEN in als je de exacte tekst van de pagina kunt
   aanwijzen waaruit je het afleidt.
2. Als je de exacte brontekst niet kunt geven: laat het veld leeg of null.
3. Verzin NOOIT iets. Geen schattingen. Geen aannames. Geen afleidingen.
4. Voor elk ingredient: geef source_text met de exacte tekst van de pagina.
5. amount is null als de hoeveelheid niet letterlijk op de pagina staat.
6. Certifications: alleen expliciete keurmerken, logos of certificeringen.
7. Geef alleen valide JSON terug zonder markdown of uitleg."""

    prompt = f"""Extraheer alle productinformatie van deze supplement-pagina.

URL: {url}

PAGINATEKST:
{full_text}

Geef terug als JSON. Laat velden leeg of null als je ze niet
letterlijk op de pagina kunt vinden:

{{
  "product_name": "exacte naam of lege string",
  "brand_name": "exacte merknaam of lege string",
  "ingredients": [
    {{
      "name": "exacte ingredientnaam zoals op label",
      "amount": 90,
      "unit": "mcg",
      "form": "exacte vorm zoals vermeld of lege string",
      "type": "actief",
      "in_proprietary_blend": false,
      "source_text": "exacte tekst van de pagina"
    }}
  ],
  "excipients": [
    {{
      "name": "exacte naam hulpstof",
      "type": "vul-additief",
      "source_text": "exacte tekst van de pagina"
    }}
  ],
  "serving_size": "exacte tekst of lege string",
  "servings_per_day": null,
  "package_size": "exacte tekst of lege string",
  "usage_instructions": "letterlijke innametekst of lege string",
  "health_claims": ["exacte claim van de pagina"],
  "certifications": ["exacte certificering zoals vermeld"],
  "price": "exacte prijs of lege string",
  "warnings": ["exacte waarschuwingstekst"],
  "batch_coa_available": false,
  "target_audience": "doelgroep als expliciet vermeld of lege string"
}}

Kritische regels:
- ingredients: alle actieve ingredienten. amount null als niet vermeld.
- excipients: capsulemateriaal, hulpstoffen, dragers. Aparte lijst.
- source_text: verplicht per ingredient. Citeer letterlijk.
- certifications: alleen als keurmerk/logo/certificering benoemd.
- batch_coa_available: true alleen bij link naar COA of labrapport."""

    try:
        response = client.messages.create(
            model=CLAUDE_MODEL,
            max_tokens=5000,
            system=system_prompt,
            messages=[{"role": "user", "content": prompt}]
        )
        raw = response.content[0].text.strip()
        raw = re.sub(r"```json\s*", "", raw)
        raw = re.sub(r"```\s*", "", raw)

        json_match = re.search(r'\{.*\}', raw, re.DOTALL)
        if not json_match:
            print("EXTRACT_CLAUDE ERROR: geen JSON object gevonden", flush=True)
            return product_data

        extracted = json.loads(json_match.group(0))
        print(f"EXTRACT_CLAUDE: {len(extracted.get('ingredients', []))} ingrediënten gevonden", flush=True)
    except Exception as e:
        print(f"EXTRACT_CLAUDE ERROR: {e}", flush=True)
        return product_data

    def merge_field(key, default=""):
        if extracted.get(key) and not product_data.get(key):
            product_data[key] = extracted[key]

    merge_field("product_name")
    merge_field("brand_name")
    merge_field("price")
    merge_field("package_size")
    merge_field("serving_size")
    merge_field("usage_instructions")
    merge_field("target_audience")

    if extracted.get("health_claims"):
        product_data["health_claims"] = extracted["health_claims"]

    if extracted.get("warnings"):
        product_data["warnings"] = extracted["warnings"]

    if extracted.get("batch_coa_available"):
        product_data["batch_coa_available"] = True

    if extracted.get("certifications"):
        existing = set(c.lower() for c in product_data.get("certifications", []))
        for cert in extracted["certifications"]:
            if cert.lower() not in existing:
                product_data.setdefault("certifications", []).append(cert)
                existing.add(cert.lower())

    if extracted.get("ingredients"):
        all_ings = extracted["ingredients"]
        existing_names = {i.get("name", "").lower() for i in all_ings}
        for ing in product_data.get("ingredients", []):
            if ing.get("name", "").lower() not in existing_names:
                all_ings.append(ing)
        active, excipients = _split_active_excipients(all_ings)
        product_data["ingredients"] = all_ings
        product_data["active_ingredients"] = active
        product_data["excipients"] = excipients

    if extracted.get("excipients") and not product_data.get("excipients"):
        product_data["excipients"] = extracted["excipients"]

    return product_data

# ═══════════════════════════════════════════════
# DEEL 7 — PRODUCT TYPE DETECTIE
# ═══════════════════════════════════════════════

def detect_product_type(product_data):
    all_text = " ".join([
        str(product_data.get("product_name", "")),
        str(product_data.get("brand_name", "")),
        " ".join(str(i.get("name", "")) for i in product_data.get("ingredients", [])),
        " ".join(product_data.get("health_claims", []))
    ]).lower()

    if any(x in all_text for x in
           ["epa", "dha", "omega-3", "omega 3", "visolie", "fish oil", "krill"]):
        return "OMEGA3"
    if any(x in all_text for x in
           ["probiotic", "lactobacillus", "bifidobacterium", "cfu"]):
        return "PROBIOTIC"
    if any(x in all_text for x in
           ["ijzer", "ferrum", "iron bisglycinate", "ferrous"]):
        return "IRON"
    if any(x in all_text for x in
           ["magnesium", "zink", "zinc", "calcium", "selenium",
            "mineraal", "mineral"]):
        return "MINERAL"
    if any(x in all_text for x in
           ["vitamine", "vitamin", "d3", "b12", "b6",
            "folaat", "folate", "b-complex"]):
        return "VITAMIN"
    if any(x in all_text for x in
           ["whey", "protein", "eiwit", "casein"]):
        return "PROTEIN"
    if any(x in all_text for x in
           ["creatine", "bcaa", "leucine", "beta-alanine", "pre-workout"]):
        return "SPORT"
    if any(x in all_text for x in
           ["curcuma", "turmeric", "berberine", "resveratrol",
            "quercetin", "polyphenol"]):
        return "BOTANICAL"
    if any(x in all_text for x in
           ["ashwagandha", "rhodiola", "valeriaan", "ginkgo", "adaptogen"]):
        return "BOTANICAL"
    return "DEFAULT"

# ═══════════════════════════════════════════════
# DEEL 8 — SCORING
# ═══════════════════════════════════════════════

def calculate_score(evaluations_data, criteria, product_type="DEFAULT"):
    weights = PRODUCT_TYPE_WEIGHTS.get(product_type, PRODUCT_TYPE_WEIGHTS["DEFAULT"])
    review_w = weights["review"]
    biology_w = weights["biology"]

    evaluations = evaluations_data.get("evaluations", [])
    core_raw = 0.0
    core_max = 0.0
    module_raw = 0.0
    module_max = 0.0
    core_critical_fail = False
    module_critical_fail = False
    non_verifiable = 0

    for i, criterion in enumerate(criteria):
        eval_item = next(
            (e for e in evaluations if e.get("criterion_index") == i + 1),
            None
        )
        pass_value = eval_item.get("pass_value", -1) if eval_item else -1
        data_quality = (
            eval_item.get("data_quality", "") if eval_item else ""
        ).upper()
        weight = criterion["weight"]
        is_core = criterion.get("source") == "core"

        if data_quality == "NIET_GEVONDEN":
            pass_value = -1

        if pass_value == -1:
            non_verifiable += 1
            continue
        elif pass_value == 1:
            if is_core:
                core_raw += weight
                core_max += weight
            else:
                module_raw += weight
                module_max += weight
        elif pass_value == 0:
            if is_core:
                core_max += weight
                if criterion["critical"] and data_quality in ("EXACT", "AFGELEID"):
                    core_critical_fail = True
            else:
                module_max += weight
                if criterion["critical"] and data_quality in ("EXACT", "AFGELEID"):
                    module_critical_fail = True

    core_pct = (core_raw / core_max) if core_max > 0 else 0.0
    module_pct = (module_raw / module_max) if module_max > 0 else 0.0

    if core_critical_fail:
        core_pct = min(core_pct, 0.49)
    if module_critical_fail:
        module_pct = min(module_pct, 0.49)

    if core_max > 0 and module_max > 0:
        overall = (core_pct * review_w) + (module_pct * biology_w)
    elif core_max > 0:
        overall = core_pct
    else:
        overall = module_pct

    critical_fail = core_critical_fail or module_critical_fail
    if critical_fail:
        overall = min(overall, 0.49)

    return overall, critical_fail, non_verifiable


def determine_verdict(score_pct, critical_fail):
    score_100 = round(score_pct * 100)
    if critical_fail and score_100 < 50:
        return score_100, "Afkeur", "Af te raden"
    if score_100 >= 85:
        return score_100, "Elite", "Koopwaardig"
    if score_100 >= 70:
        return score_100, "Degelijk", "Koopwaardig"
    if score_100 >= 50:
        return score_100, "Matig", "Alleen met context"
    return score_100, "Afkeur", "Af te raden"

# ═══════════════════════════════════════════════
# DEEL 9 — CRITERIA EVALUATIE MET CLAUDE
# ═══════════════════════════════════════════════

def evaluate_criteria_with_claude(product_data, criteria, product_type, url=""):
    client = anthropic.Anthropic(api_key=os.environ.get("ANTHROPIC_API_KEY"))

    # Filter relevante criteria
    relevant = []
    for c in criteria:
        applies = c["applies_to"]
        if (applies == "ALL" or applies == product_type or
                (applies == "SPECIFIC" and product_type != "DEFAULT")):
            relevant.append(c)

    core_criteria = [c for c in relevant if c["source"] == "core"]
    module_criteria = [c for c in relevant if c["source"] == "module"]

    if not core_criteria and not module_criteria:
        print("GEEN CRITERIA BESCHIKBAAR", flush=True)
        return {"evaluations": [], "key_strengths": [],
                "key_weaknesses": [], "inferior_forms_found": []}, []

    # Gedeelde productcontext
    ingredients_text = "\n".join([
        f"- {i.get('name','')}: {i.get('amount','?')} "
        f"{i.get('unit','')} (vorm: {i.get('form','niet vermeld')})"
        for i in product_data.get("ingredients", [])[:30]
        if i.get("type") != "vul-additief"
    ])

    product_context = f"""Product: {product_data.get('product_name','Onbekend')}
Merk: {product_data.get('brand_name','Onbekend')}
Type: {product_type}
Serving size: {product_data.get('serving_size','onbekend')}
Verpakking: {product_data.get('package_size','onbekend')}
Prijs: {product_data.get('price','onbekend')}

ACTIEVE INGREDIENTEN:
{ingredients_text or 'Geen ingredienten gevonden'}

EXCIPIENTS:
{chr(10).join(str(e.get('name','')) for e in product_data.get('excipients',[])[:20]) or 'Geen'}

CERTIFICERINGEN:
{', '.join(product_data.get('certifications',[])) or 'Geen'}

GEZONDHEIDSCLAIMS:
{chr(10).join(product_data.get('health_claims',[])[:8]) or 'Geen'}

AANVULLENDE INFO:
{product_data.get('additional_info','')[:500]}"""

    system_prompt = """Je bent de Funcify scoring engine.

SCORING REGELS — absoluut:
- Pass=1: bewijs aanwezig EN criterium gehaald
- Pass=0: bewijs aanwezig EN criterium NIET gehaald
- Pass=-1: informatie niet beschikbaar — score-neutraal
- Pass=-1 activeert NOOIT een Critical Gate
- Alleen Pass=0 met EXACT of AFGELEID confidence activeert Critical Gate
- Ontbrekende data is ALTIJD Pass=-1, nooit Pass=0

DATA QUALITY:
- EXACT: letterlijk op label of pagina vermeld
- AFGELEID: logisch af te leiden
- ONVOLLEDIG: deels beschikbaar
- NIET_GEVONDEN: niet aangetroffen

Geef alleen valide JSON terug zonder markdown."""

    all_evaluations = []
    all_strengths = []
    all_weaknesses = []
    all_inferior = []
    offset = 0

    # Call 1: core criteria
    if core_criteria:
        criteria_text = "\n".join([
            f"{i+1}. [{c['category'][:30]}] {c['criterion'][:80]} "
            f"(weight={c['weight']}, critical={'JA' if c['critical'] else 'NEE'})"
            for i, c in enumerate(core_criteria)
        ])

        user_prompt = f"""{product_context}

CRITERIA TE EVALUEREN (CORE — {len(core_criteria)} stuks):
{criteria_text}

Geef terug als JSON:
{{
  "evaluations": [
    {{
      "criterion_index": 1,
      "pass_value": 1,
      "evidence": "korte uitleg max 1 zin",
      "data_quality": "EXACT|AFGELEID|ONVOLLEDIG|NIET_GEVONDEN"
    }}
  ],
  "key_strengths": ["max 3 sterke punten in gewone taal"],
  "key_weaknesses": ["max 3 zwakke punten in gewone taal"],
  "inferior_forms_found": ["inferieure vormen indien aanwezig"]
}}"""

        try:
            response = client.messages.create(
                model=CLAUDE_MODEL,
                max_tokens=6000,
                system=system_prompt,
                messages=[{"role": "user", "content": user_prompt}]
            )
            raw = response.content[0].text.strip()
            raw = re.sub(r"```json\s*", "", raw)
            raw = re.sub(r"```\s*", "", raw)
            json_match = re.search(r'\{.*\}', raw, re.DOTALL)
            if json_match:
                result = json.loads(json_match.group(0))
                evals = result.get("evaluations", [])
                all_evaluations.extend(evals)
                all_strengths = result.get("key_strengths", [])
                all_weaknesses = result.get("key_weaknesses", [])
                all_inferior = result.get("inferior_forms_found", [])
                offset = len(core_criteria)
                print(f"CORE EVALUATIE: {len(evals)} criteria gescoord", flush=True)
        except Exception as e:
            print(f"CORE EVALUATIE ERROR: {e}", flush=True)

    # Call 2: module criteria
    if module_criteria:
        criteria_text = "\n".join([
            f"{offset+i+1}. [{c['category'][:30]}] {c['criterion'][:80]} "
            f"(weight={c['weight']}, critical={'JA' if c['critical'] else 'NEE'})"
            for i, c in enumerate(module_criteria)
        ])

        user_prompt = f"""{product_context}

CRITERIA TE EVALUEREN (MODULE {product_type} — {len(module_criteria)} stuks):
{criteria_text}

Geef terug als JSON:
{{
  "evaluations": [
    {{
      "criterion_index": {offset+1},
      "pass_value": 1,
      "evidence": "korte uitleg max 1 zin",
      "data_quality": "EXACT|AFGELEID|ONVOLLEDIG|NIET_GEVONDEN"
    }}
  ],
  "key_strengths": ["aanvullende sterke punten specifiek voor {product_type}"],
  "key_weaknesses": ["aanvullende zwakke punten specifiek voor {product_type}"],
  "inferior_forms_found": []
}}"""

        try:
            response = client.messages.create(
                model=CLAUDE_MODEL,
                max_tokens=4000,
                system=system_prompt,
                messages=[{"role": "user", "content": user_prompt}]
            )
            raw = response.content[0].text.strip()
            raw = re.sub(r"```json\s*", "", raw)
            raw = re.sub(r"```\s*", "", raw)
            json_match = re.search(r'\{.*\}', raw, re.DOTALL)
            if json_match:
                result = json.loads(json_match.group(0))
                evals = result.get("evaluations", [])
                all_evaluations.extend(evals)
                all_strengths += result.get("key_strengths", [])
                all_weaknesses += result.get("key_weaknesses", [])
                print(f"MODULE EVALUATIE: {len(evals)} criteria gescoord", flush=True)
        except Exception as e:
            print(f"MODULE EVALUATIE ERROR: {e}", flush=True)

    all_relevant = core_criteria + module_criteria
    print(f"TOTAAL EVALUATIE: {len(all_evaluations)} van {len(all_relevant)} criteria", flush=True)

    return {
        "evaluations": all_evaluations,
        "key_strengths": all_strengths[:3],
        "key_weaknesses": all_weaknesses[:3],
        "inferior_forms_found": all_inferior
    }, all_relevant

# ═══════════════════════════════════════════════
# DEEL 10 — CONSUMER OUTPUT GENERATIE
# ═══════════════════════════════════════════════

def generate_consumer_output(product_data, evaluations_data, criteria,
                              score_100, kwalificatie, verdict,
                              product_type, critical_fail,
                              context_flags_triggered=None):
    client = anthropic.Anthropic(api_key=os.environ.get("ANTHROPIC_API_KEY"))

    all_certs = product_data.get("certifications", [])
    quality_certs = [
        c for c in all_certs
        if any(q in str(c).lower() for q in QUALITY_CERT_LIST)
    ]
    sustainability_certs = [
        c for c in all_certs
        if any(s in str(c).lower() for s in SUSTAINABILITY_CERT_LIST)
    ]

    if quality_certs and sustainability_certs:
        cert_oordeel = "GOED"
    elif quality_certs or sustainability_certs:
        cert_oordeel = "MATIG"
    else:
        cert_oordeel = "ONTBREEKT"

    spd = _parse_servings_per_day(
        product_data.get("serving_size", ""),
        product_data.get("usage_instructions", "")
    )
    portie_tekst = (
        "PORTIEGROOTTE: meer dan 4 eenheden per dag. Rij 7 = MATIG."
        if spd > 4 else ""
    )

    if score_100 >= 75:
        s6_instructie = (
            "VOOR WIE IS DIT PRODUCT GESCHIKT: schrijf 2-3 zinnen "
            "in tweede persoon over wie dit product het meest geschikt is. "
            "Geen verbeterpunten."
        )
        s6_key = "voor_wie_geschikt"
        s7_tonen = False
    elif score_100 >= 50:
        s6_instructie = (
            "WAT ZOU EEN BETER PRODUCT BEVATTEN: maximaal 2 "
            "verbeterpunten plus korte doelgroepomschrijving. "
            "Maximaal 3 zinnen."
        )
        s6_key = "wat_zou_beter"
        s7_tonen = True
    else:
        s6_instructie = (
            "WAT ZOU EEN BETER PRODUCT BEVATTEN: alleen verbeterpunten. "
            "Sluit af met: Vraag bij aankoop altijd om een onafhankelijk "
            "testrapport per batch."
        )
        s6_key = "wat_zou_beter"
        s7_tonen = True

    usage = str(product_data.get("usage_instructions") or "").strip()
    inname = (
        f"INNAME-ADVIES VAN LABEL: {usage}"
        if usage and len(usage) > 10
        else "Geen inname-advies gevonden op label — NIET vermelden."
    )

    ui_content = _cache.get("ui_content", {})
    unknown_ingredients = []
    for ing in product_data.get("ingredients", []):
        if ing.get("type") == "vul-additief":
            continue
        name = str(ing.get("name", "")).lower()
        found = any(
            name in str(v.get("display_name", "")).lower()
            for v in ui_content.values()
        )
        if not found and name:
            unknown_ingredients.append(ing.get("name", ""))

    flags_tekst = "\n".join([
        f.get("message", "")
        for f in (context_flags_triggered or [])[:6]
    ])

    critical_gate_str = "JA" if critical_fail else "NEE"
    s7_str = "JA — tonen" if s7_tonen else "NEE — weglaten"

    system_prompt = """Je bent de Funcify beoordelingsengine voor Nederlandse consumenten.

UNIVERSELE REGELS — nooit overtreden:
1. Altijd tweede persoon (je, jij). Nooit derde persoon.
2. Maximaal 2-3 zinnen per tabelrij in kolom Wat we zien.
3. VERBODEN: TG, rTG, COA, TOTOX, CFU, UL, EFSA, HPLC, ING-codes,
   DATA_LACUNE, DB-codes, therapeutisch, klinisch bewezen.
4. Oordelen: uitsluitend GOED, MATIG, SLECHT of ONTBREEKT.
5. ONTBREEKT = info niet beschikbaar. SLECHT = bewezen slecht.
6. Inname-advies ALLEEN als letterlijk op label staat.
7. Rij 8: ALLEEN excipients lijst gebruiken. Nooit aannames.
8. Minimaal 2 sterktes EN minimaal 2 zwaktes in highlights.
9. Goed nieuws — als prefix voor sterktes.
10. Let op — als prefix voor zwaktes.
11. Geef alleen valide JSON terug zonder markdown."""

    user_prompt = f"""Product: {product_data.get('product_name', 'Onbekend')} \
({product_data.get('brand_name', 'Onbekend')})
Score: {score_100}/100 | {kwalificatie} | {verdict}
Producttype: {product_type} | Critical gate: {critical_gate_str}

ACTIEVE INGREDIENTEN:
{chr(10).join([
    f"- {i.get('name', '')}: {i.get('amount', '')} {i.get('unit', '')} "
    f"(vorm: {i.get('form', 'niet vermeld')})"
    for i in product_data.get('ingredients', [])
    if i.get('type') != 'vul-additief'
]) or 'Geen ingredienten gevonden'}

EXCIPIENTS (ALLEEN voor rij 8):
{chr(10).join([
    str(e.get('name', e) if isinstance(e, dict) else e)
    for e in product_data.get('excipients', [])
]) or 'Geen excipients — rij 8 = ONTBREEKT'}

{inname}

GEZONDHEIDSCLAIMS:
{chr(10).join(product_data.get('health_claims', [])[:8]) or 'Geen'}

KWALITEITSCERTIFICERINGEN (rij 5 categorie 1):
{', '.join(quality_certs) or 'Geen'}

DUURZAAMHEIDSCERTIFICERINGEN (rij 5 categorie 2):
{', '.join(sustainability_certs) or 'Geen'}

VOORBEREKEND OORDEEL RIJ 5: {cert_oordeel}
{portie_tekst}

STERKE PUNTEN: {', '.join(evaluations_data.get('key_strengths', []))}
ZWAKKE PUNTEN: {', '.join(evaluations_data.get('key_weaknesses', []))}
CONTEXT WAARSCHUWINGEN: {flags_tekst or 'Geen'}

SECTIE 6: {s6_instructie}
SECTIE 7 TONEN: {s7_str}

Genereer JSON:
{{
  "wat_doet": "2-3 zinnen. Wat doet het supplement. Voor wie relevant.",
  "beoordeling_tabel": [
    {{"aspect": "Welke vorm krijg je?", "bevinding": "...", "oordeel": "GOED|MATIG|SLECHT|ONTBREEKT"}},
    {{"aspect": "Krijg je genoeg per dag?", "bevinding": "...", "oordeel": "GOED|MATIG|SLECHT|ONTBREEKT"}},
    {{"aspect": "Hoe goed wordt het opgenomen?", "bevinding": "...", "oordeel": "GOED|MATIG|SLECHT|ONTBREEKT"}},
    {{"aspect": "Is alles duidelijk vermeld?", "bevinding": "...", "oordeel": "GOED|MATIG|SLECHT|ONTBREEKT"}},
    {{"aspect": "Onafhankelijk gecontroleerd?", "bevinding": "kwaliteitscert: {', '.join(quality_certs) or 'geen'} | duurzaamheidscert: {', '.join(sustainability_certs) or 'geen'}", "oordeel": "{cert_oordeel}"}},
    {{"aspect": "Kloppen de gezondheidsclaims?", "bevinding": "...", "oordeel": "GOED|MATIG|SLECHT|ONTBREEKT"}},
    {{"aspect": "Hoeveel per verpakking?", "bevinding": "...", "oordeel": "GOED|MATIG|SLECHT|ONTBREEKT"}},
    {{"aspect": "Zijn er onnodige toevoegingen?", "bevinding": "...", "oordeel": "GOED|MATIG|SLECHT|ONTBREEKT"}}
  ],
  "highlights": [
    {{"type": "positief", "tekst": "Goed nieuws — ..."}},
    {{"type": "positief", "tekst": "Goed nieuws — ..."}},
    {{"type": "negatief", "tekst": "Let op — ..."}},
    {{"type": "negatief", "tekst": "Let op — ..."}}
  ],
  "{s6_key}": "tekst conform sectie 6 instructie",
  "voor_wie": "",
  "consumer_summary": "één zin max 30 woorden met sterkste en zwakste punt"
}}"""

    response = client.messages.create(
        model=CLAUDE_MODEL,
        max_tokens=5000,
        system=system_prompt,
        messages=[{"role": "user", "content": user_prompt}]
    )
    raw = response.content[0].text.strip()
    raw = re.sub(r"```json\s*", "", raw)
    raw = re.sub(r"```\s*", "", raw)
    try:
        json_match = re.search(r'\{.*\}', raw, re.DOTALL)
        if not json_match:
            print("CONSUMER_OUTPUT ERROR: geen JSON object gevonden", flush=True)
            raise ValueError("Geen JSON object in response")
        result = json.loads(json_match.group(0))
    except json.JSONDecodeError as e:
        print(f"CONSUMER_OUTPUT JSON ERROR: {e}", flush=True)
        raise

    if unknown_ingredients:
        result.setdefault("highlights", []).append({
            "type": "info",
            "tekst": f"Ingrediënt niet herkend — Funcify gaat dit onderzoeken: {', '.join(unknown_ingredients[:3])}"
        })

    result.setdefault("voor_wie_geschikt", "")
    result.setdefault("wat_zou_beter", "")
    result.setdefault("voor_wie", "")
    result["vaste_afsluiting"] = VASTE_AFSLUITING

    return result

# ═══════════════════════════════════════════════
# DEEL 11 — ROUTES
# ═══════════════════════════════════════════════

@app.route("/health", methods=["GET"])
def health():
    criteria_count = len(_cache.get("criteria", []))
    ui_count = len(_cache.get("ui_content", {}))
    return jsonify({
        "status": "ok",
        "criteria_loaded": criteria_count,
        "ui_content_loaded": ui_count,
        "scrapingbee": bool(os.environ.get("SCRAPINGBEE_KEY"))
    }), 200


@app.route("/test-scrapingbee", methods=["GET"])
def test_scrapingbee():
    import time
    api_key = os.environ.get("SCRAPINGBEE_KEY")
    if not api_key:
        print("TEST_SCRAPINGBEE: geen API key gevonden", flush=True)
        return jsonify({"error": "geen SCRAPINGBEE_KEY"}), 500

    test_url = "https://example.com"
    print(f"TEST_SCRAPINGBEE: verzoek naar {test_url}", flush=True)
    start = time.time()
    try:
        response = requests.get(
            "https://app.scrapingbee.com/api/v1/",
            params={
                "api_key": api_key,
                "url": test_url,
                "render_js": "true",
                "wait": "1000",
                "block_resources": "true",
            },
            timeout=30,
        )
        elapsed_ms = round((time.time() - start) * 1000)
        print(f"TEST_SCRAPINGBEE: status={response.status_code} tijd={elapsed_ms}ms", flush=True)
        return jsonify({
            "status_code": response.status_code,
            "elapsed_ms": elapsed_ms,
            "response_size_bytes": len(response.content),
            "ok": response.ok,
        })
    except requests.Timeout:
        elapsed_ms = round((time.time() - start) * 1000)
        print(f"TEST_SCRAPINGBEE: TIMEOUT na {elapsed_ms}ms", flush=True)
        return jsonify({"error": "timeout", "elapsed_ms": elapsed_ms}), 504
    except Exception as e:
        elapsed_ms = round((time.time() - start) * 1000)
        print(f"TEST_SCRAPINGBEE: FOUT na {elapsed_ms}ms — {e}", flush=True)
        return jsonify({"error": str(e), "elapsed_ms": elapsed_ms}), 500


def scrape_only(url):
    """
    Scraping pipeline — geeft product_data terug.
    Scoring wordt gedaan door de /score route.
    """
    product_data = {}
    page_text = ""

    # Laag 1: ophalen
    try:
        soup, page_text = fetch_with_scrapingbee(url)
        product_data = scrape_product(soup, url)
        product_data["_source"] = "scraper"
        print(f"WORKER SCRAPE: {len(page_text)} tekens opgehaald", flush=True)
    except Exception as e:
        print(f"WORKER SCRAPE ERROR: {e}", flush=True)
        product_data["_source"] = "error"

    # Laag 2: Claude parser
    try:
        product_data = extract_with_claude(url, page_text, product_data)
        print(f"WORKER EXTRACT: {len(product_data.get('ingredients', []))} ingrediënten", flush=True)
    except Exception as e:
        print(f"WORKER EXTRACT ERROR: {e}", flush=True)

    # Verdedigingslaag
    product_data["product_name"] = str(product_data.get("product_name") or "Onbekend")
    product_data["brand_name"] = str(product_data.get("brand_name") or "Onbekend")
    product_data["ingredients"] = product_data.get("ingredients") or []
    product_data["active_ingredients"] = product_data.get("active_ingredients") or []
    product_data["excipients"] = product_data.get("excipients") or []
    product_data["health_claims"] = product_data.get("health_claims") or []
    product_data["certifications"] = product_data.get("certifications") or []
    product_data["warnings"] = product_data.get("warnings") or []
    product_data["serving_size"] = str(product_data.get("serving_size") or "")
    product_data["usage_instructions"] = str(product_data.get("usage_instructions") or "")
    product_data["package_size"] = str(product_data.get("package_size") or "")
    product_data["price"] = str(product_data.get("price") or "")
    product_data["additional_info"] = str(product_data.get("additional_info") or "")
    product_data["problem_ids"] = product_data.get("problem_ids") or []
    product_data["url"] = url

    for ing in product_data["ingredients"]:
        ing["name"] = str(ing.get("name") or "")
        ing["form"] = str(ing.get("form") or "")
        ing["unit"] = str(ing.get("unit") or "")
        ing["type"] = str(ing.get("type") or "actief")
        if ing.get("amount") is not None:
            try:
                ing["amount"] = float(ing["amount"])
            except (ValueError, TypeError):
                ing["amount"] = None

    return product_data


@app.route("/score", methods=["POST"])
def score():
    product_data = request.get_json(silent=True) or {}

    # Verdedigingslaag
    product_data["product_name"] = str(product_data.get("product_name") or "Onbekend")
    product_data["brand_name"] = str(product_data.get("brand_name") or "Onbekend")
    product_data["ingredients"] = product_data.get("ingredients") or []
    product_data["active_ingredients"] = product_data.get("active_ingredients") or []
    product_data["excipients"] = product_data.get("excipients") or []
    product_data["health_claims"] = product_data.get("health_claims") or []
    product_data["certifications"] = product_data.get("certifications") or []
    product_data["serving_size"] = str(product_data.get("serving_size") or "")
    product_data["usage_instructions"] = str(product_data.get("usage_instructions") or "")
    product_data["package_size"] = str(product_data.get("package_size") or "")
    product_data["price"] = str(product_data.get("price") or "")
    product_data["additional_info"] = str(product_data.get("additional_info") or "")
    product_data["warnings"] = product_data.get("warnings") or []
    product_data["problem_ids"] = product_data.get("problem_ids") or []

    for ing in product_data["ingredients"]:
        ing["name"] = str(ing.get("name") or "")
        ing["form"] = str(ing.get("form") or "")
        ing["unit"] = str(ing.get("unit") or "")
        ing["type"] = str(ing.get("type") or "actief")
        if ing.get("amount") is not None:
            try:
                ing["amount"] = float(ing["amount"])
            except (ValueError, TypeError):
                ing["amount"] = None

    url = str(product_data.get("url") or "")
    product_data["url"] = url

    # Stap 1: producttype
    try:
        product_type = detect_product_type(product_data)
        print(f"STAP1: producttype = {product_type}", flush=True)
    except Exception as e:
        print(f"STAP1 ERROR: {e}", flush=True)
        product_type = "DEFAULT"

    # Stap 2: complexiteit en prijs
    try:
        active_ings = product_data.get("active_ingredients") or [
            i for i in product_data.get("ingredients", [])
            if i.get("type") != "vul-additief"
        ]
        active_count = len(active_ings)
        if active_count <= 2:
            tier = "Single"
        elif active_count <= 6:
            tier = "Multi"
        else:
            tier = "Complex"
        price_per_day = calculate_price_per_day(
            product_data.get("price", ""),
            product_data.get("package_size", ""),
            product_data.get("serving_size", ""),
            product_data.get("usage_instructions", "")
        )
        price_per_gram = calculate_price_per_gram(
            product_data.get("price", ""),
            product_data.get("ingredients", []),
            product_data.get("package_size", ""),
            product_data.get("serving_size", "")
        )
        intake_advice = get_intake_advice(product_data)
        print(f"STAP2: tier={tier}, ppd={price_per_day}", flush=True)
    except Exception as e:
        print(f"STAP2 ERROR: {e}", flush=True)
        tier = "Single"
        price_per_day = ""
        price_per_gram = ""
        intake_advice = ""

    # Stap 3: context flags
    try:
        active_ings = product_data.get("active_ingredients") or [
            i for i in product_data.get("ingredients", [])
            if i.get("type") != "vul-additief"
        ]
        flags = evaluate_context_flags(active_ings)
        flags += evaluate_cofactor_checks(active_ings)
        flags.sort(key=lambda f: SEVERITY_ORDER.get(f.get("severity", "Info"), 4))
        print(f"STAP3: {len(flags)} flags", flush=True)
    except Exception as e:
        print(f"STAP3 ERROR: {e}", flush=True)
        flags = []

    # Stap 4: criteria evaluatie
    try:
        criteria = _cache.get("criteria", [])
        print(f"STAP4: {len(criteria)} criteria geladen uit cache", flush=True)
        eval_data, relevant_criteria = evaluate_criteria_with_claude(
            product_data, criteria, product_type, url
        )
    except Exception as e:
        print(f"STAP4 ERROR: {e}", flush=True)
        eval_data = {
            "evaluations": [], "key_strengths": [],
            "key_weaknesses": [], "inferior_forms_found": []
        }
        relevant_criteria = []

    # Stap 5: score
    try:
        score_pct, critical_fail, non_verifiable = calculate_score(
            eval_data, relevant_criteria, product_type
        )
        score_100, kwalificatie, verdict = determine_verdict(
            score_pct, critical_fail
        )
        evals = eval_data.get("evaluations", [])
        exact_count = sum(
            1 for e in evals
            if e.get("data_quality", "").upper() == "EXACT"
        )
        confidence = exact_count / len(evals) if evals else 0.0
        print(f"STAP5: score={score_100}, critical={critical_fail}", flush=True)
    except Exception as e:
        print(f"STAP5 ERROR: {e}", flush=True)
        score_100 = 0
        kwalificatie = "Onbekend"
        verdict = "Niet beschikbaar"
        critical_fail = False
        non_verifiable = 0
        confidence = 0.0

    # Stap 6: consumer output
    try:
        output = generate_consumer_output(
            product_data, eval_data, relevant_criteria,
            score_100, kwalificatie, verdict, product_type,
            critical_fail, context_flags_triggered=flags
        )
    except Exception as e:
        print(f"STAP6 ERROR: {e}", flush=True)
        output = {
            "wat_doet": "Beoordeling tijdelijk niet beschikbaar.",
            "beoordeling_tabel": [],
            "highlights": [],
            "voor_wie_geschikt": "",
            "wat_zou_beter": "",
            "voor_wie": "",
            "consumer_summary": "Beoordeling kon niet worden gegenereerd.",
            "vaste_afsluiting": VASTE_AFSLUITING,
        }

    # Stap 7: jargon
    try:
        output = simplify_jargon(output)
    except Exception as e:
        print(f"STAP7 ERROR: {e}", flush=True)

    # Certificeringen splitsen
    try:
        all_certs = product_data.get("certifications", [])
        q_certs = [c for c in all_certs if any(q in str(c).lower() for q in QUALITY_CERT_LIST)]
        s_certs = [c for c in all_certs if any(s in str(c).lower() for s in SUSTAINABILITY_CERT_LIST)]
    except Exception as e:
        print(f"CERTS ERROR: {e}", flush=True)
        q_certs = []
        s_certs = []

    response_body = {
        "product_name": product_data.get("product_name", "Onbekend"),
        "brand": product_data.get("brand_name", "Onbekend"),
        "brand_name": product_data.get("brand_name", "Onbekend"),
        "score": score_100,
        "kwalificatie": kwalificatie,
        "verdict": verdict,
        "product_type": product_type,
        "product_complexity_tier": tier,
        "product_complexity_label": COMPLEXITY_LABELS.get(tier, tier),
        "critical_gate": critical_fail,
        "non_verifiable_count": non_verifiable,
        "criteria_evaluated": len(relevant_criteria),
        "confidence": round(confidence, 2),
        "price": product_data.get("price", ""),
        "package_size": product_data.get("package_size", ""),
        "price_per_day": price_per_day,
        "price_per_gram": price_per_gram,
        "intake_advice": intake_advice,
        "context_flags": flags,
        "context_flags_triggered": flags,
        "quality_certs": q_certs,
        "sustainability_certs": s_certs,
        "problem_ids": product_data.get("problem_ids", []),
        **output
    }

    return jsonify(response_body)


@app.route("/scrape", methods=["POST"])
def scrape():
    data = request.get_json(silent=True) or {}
    url = str(data.get("url", "")).strip()
    if not url:
        return jsonify({"error": "URL required"}), 400

    # Probeer Redis queue
    conn = get_redis_connection()
    if conn and HAS_REDIS:
        try:
            q = Queue(connection=conn, default_timeout=600)
            job = q.enqueue(scrape_only, url)
            print(f"QUEUE: job {job.id} aangemaakt voor {url}", flush=True)
            return jsonify({
                "job_id": job.id,
                "status": "queued"
            })
        except Exception as e:
            print(f"QUEUE ERROR: {e}, fallback naar synchroon", flush=True)

    # Fallback: synchrone verwerking als Redis niet beschikbaar is
    print("SCRAPE: synchrone verwerking (geen Redis)", flush=True)
    result = scrape_only(url)
    result["job_id"] = "sync"
    result["status"] = "completed"
    return jsonify(result)


@app.route("/result/<job_id>", methods=["GET"])
def get_result(job_id):
    """Poll endpoint voor job resultaat."""

    # Sync fallback had geen echte job_id
    if job_id == "sync":
        return jsonify({"status": "completed"})

    conn = get_redis_connection()
    if not conn or not HAS_REDIS:
        return jsonify({"status": "error", "error": "Redis niet beschikbaar"}), 500

    try:
        job = Job.fetch(job_id, connection=conn)
        if job.is_finished:
            result = job.result.copy()
            result["status"] = "completed"
            return jsonify(result)
        elif job.is_failed:
            return jsonify({"status": "failed", "error": str(job.exc_info)}), 500
        else:
            return jsonify({"status": "processing"})
    except Exception as e:
        return jsonify({"status": "error", "error": str(e)}), 500


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 8080))
    app.run(host="0.0.0.0", port=port)
